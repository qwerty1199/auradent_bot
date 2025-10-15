#!/usr/bin/env python3
"""
AuraDent Bot - Telegram bot for handling dental consultation requests
"""

import os
import logging
import json
from datetime import datetime
from typing import Dict, Any, Optional
from pathlib import Path

from telegram import Update, ForceReply
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    filters
)
from dotenv import load_dotenv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Load environment variables
load_dotenv()

# Configuration
BOT_TOKEN = os.getenv('BOT_TOKEN')
ADMIN_CHAT_ID = os.getenv('ADMIN_CHAT_ID')
EXCEL_FILE_PATH = os.getenv('EXCEL_FILE_PATH', 'consultations.xlsx')
LOG_LEVEL = os.getenv('LOG_LEVEL', 'INFO')

# Webhook configuration for cloud hosting
WEBHOOK_URL = os.getenv('WEBHOOK_URL')
PORT = int(os.getenv('PORT', 8443))
WEBHOOK_PATH = f"/{BOT_TOKEN}"

# Setup logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=getattr(logging, LOG_LEVEL.upper())
)
logger = logging.getLogger(__name__)


class ConsultationManager:
    """Manages consultation requests and Excel operations"""
    
    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        self._ensure_excel_file_exists()
    
    def _ensure_excel_file_exists(self) -> None:
        """Create Excel file with headers if it doesn't exist"""
        if not self.excel_path.exists():
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Consultations"
            
            # Define headers
            headers = [
                "Timestamp", "Name", "Phone", "Email", 
                "Age", "Consultation Type", "Message", "Chat ID"
            ]
            
            # Add headers with formatting
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(self.excel_path)
            logger.info(f"Created Excel file: {self.excel_path}")
    
    def add_consultation(self, consultation_data: Dict[str, Any]) -> bool:
        """Add a new consultation to the Excel file"""
        try:
            workbook = openpyxl.load_workbook(self.excel_path)
            sheet = workbook.active
            
            # Find next empty row
            next_row = sheet.max_row + 1
            
            # Add timestamp
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Map data to columns
            row_data = [
                timestamp,
                consultation_data.get('name', ''),
                consultation_data.get('phone', ''),
                consultation_data.get('email', ''),
                consultation_data.get('age', ''),
                consultation_data.get('consultation_type', ''),
                consultation_data.get('message', ''),
                consultation_data.get('chat_id', '')
            ]
            
            # Add data to row
            for col, value in enumerate(row_data, 1):
                sheet.cell(row=next_row, column=col, value=value)
            
            workbook.save(self.excel_path)
            logger.info(f"Added consultation to Excel: {consultation_data.get('name', 'Unknown')}")
            return True
            
        except Exception as e:
            logger.error(f"Error adding consultation to Excel: {e}")
            return False
    
    def get_excel_file_path(self) -> Path:
        """Get the path to the Excel file"""
        return self.excel_path
    
    def get_consultation_count(self) -> int:
        """Get the total number of consultations"""
        try:
            workbook = openpyxl.load_workbook(self.excel_path)
            sheet = workbook.active
            return sheet.max_row - 1  # Subtract 1 for header row
        except Exception as e:
            logger.error(f"Error counting consultations: {e}")
            return 0


# Initialize consultation manager
consultation_manager = ConsultationManager(EXCEL_FILE_PATH)


def parse_consultation_message(message_text: str) -> Optional[Dict[str, Any]]:
    """
    Parse consultation request from message text
    Expected format: 
    New Consultation Request
    Name: Patient Name
    Email: patient@email.com
    Phone: 1234567890
    Message: Optional message (optional)
    Date: Optional date (optional)
    """
    try:
        # Try to parse as JSON first
        if message_text.strip().startswith('{'):
            return json.loads(message_text)
        
        # Check if it starts with "New Consultation Request"
        lines = message_text.strip().split('\n')
        
        # Skip the first line if it's "New Consultation Request"
        start_index = 0
        if lines and lines[0].strip().lower() in ['new consultation request', 'consultation request']:
            start_index = 1
        
        consultation_data = {}
        
        # Parse key-value pairs
        for i in range(start_index, len(lines)):
            line = lines[i].strip()
            if ':' in line:
                key, value = line.split(':', 1)
                key = key.strip().lower()
                value = value.strip()
                
                # Skip empty values
                if not value:
                    continue
                
                # Map keys to standardized names
                key_mapping = {
                    'name': 'name',
                    'email': 'email', 
                    'phone': 'phone',
                    'message': 'message',
                    'date': 'date',
                    'consultation type': 'consultation_type',
                    'age': 'age'
                }
                
                mapped_key = key_mapping.get(key, key.replace(' ', '_'))
                consultation_data[mapped_key] = value
        
        # Validate required fields (name, email, phone)
        required_fields = ['name', 'email', 'phone']
        if not all(field in consultation_data for field in required_fields):
            logger.warning(f"Missing required fields in consultation: {consultation_data}")
            return None
        
        return consultation_data if consultation_data else None
        
    except Exception as e:
        logger.error(f"Error parsing consultation message: {e}")
        return None


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Send a message when the command /start is issued."""
    user = update.effective_user
    await update.message.reply_html(
        f"Hi {user.mention_html()}!\n\n"
        f"🦷 Welcome to AuraDent Bot!\n\n"
        f"This bot handles dental consultation requests. "
        f"Send consultation data in the following format:\n\n"
        f"<code>New Consultation Request\n"
        f"Name: Eugeniu Buzila\n"
        f"Email: eugeniubuzila11@gmail.com\n"
        f"Phone: 3886356363\n"
        f"Message: Your message here (optional)\n"
        f"Date: Sat Oct 18 2025 (optional)</code>\n\n"
        f"Admins can use /get_consultations to download the Excel file.",
        reply_markup=ForceReply(selective=True),
    )


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Send a message when the command /help is issued."""
    help_text = """
🦷 *AuraDent Bot Help*

*Available Commands:*
/start - Start the bot and see instructions
/help - Show this help message
/myid - Show your user ID for admin setup
/get_consultations - Download Excel file with all consultations (admin only)
/stats - Show consultation statistics

*For Consultation Requests:*
Send a message with consultation details in this format:

```
New Consultation Request
Name: Eugeniu Buzila
Email: eugeniubuzila11@gmail.com
Phone: 3886356363
Message: Your message here (optional)
Date: Sat Oct 18 2025 (optional)
```

*Required fields:* Name, Email, Phone
*Optional fields:* Message, Date

*JSON Format is also supported:*
```json
{
    "name": "Patient Name",
    "email": "patient@email.com",
    "phone": "1234567890",
    "message": "Optional message",
    "date": "Optional date"
}
```
    """
    await update.message.reply_text(help_text, parse_mode='Markdown')


async def get_consultations(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Send the Excel file with all consultations (admin only)."""
    chat_id = str(update.effective_chat.id)
    user_id = str(update.effective_user.id)
    
    # Debug info
    logger.info(f"get_consultations called - Chat ID: {chat_id}, User ID: {user_id}, Admin ID: {ADMIN_CHAT_ID}")
    
    # Check if user is admin (check both user ID and chat ID for flexibility)
    if ADMIN_CHAT_ID and user_id != ADMIN_CHAT_ID and chat_id != ADMIN_CHAT_ID:
        await update.message.reply_text(
            f"❌ Access denied. This command is for authorized users only.\n"
            f"Your User ID: {user_id}\n"
            f"Chat ID: {chat_id}\n"
            f"Use /myid in a private chat with the bot to get your correct ID."
        )
        return
    
    # If ADMIN_CHAT_ID is not set, show setup instructions
    if not ADMIN_CHAT_ID:
        await update.message.reply_text(
            f"⚙️ Admin access not configured.\n"
            f"Your User ID: {user_id}\n"
            f"Chat ID: {chat_id}\n\n"
            f"Set ADMIN_CHAT_ID={user_id} in your environment variables."
        )
        return
    
    try:
        excel_path = consultation_manager.get_excel_file_path()
        
        if not excel_path.exists():
            await update.message.reply_text("📋 No consultations file found.")
            return
        
        # Get consultation count
        count = consultation_manager.get_consultation_count()
        logger.info(f"Excel file exists: {excel_path.exists()}, Consultation count: {count}")
        
        # Send the Excel file
        with open(excel_path, 'rb') as file:
            await update.message.reply_document(
                document=file,
                filename=f"consultations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                caption=f"📊 Consultations Export\n"
                       f"Total consultations: {count}\n"
                       f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            )
        
        logger.info(f"Excel file sent to admin (Chat ID: {chat_id})")
        
    except Exception as e:
        logger.error(f"Error sending Excel file: {e}")
        await update.message.reply_text("❌ Error generating consultations file. Please try again later.")


async def stats_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Show consultation statistics."""
    try:
        count = consultation_manager.get_consultation_count()
        excel_path = consultation_manager.get_excel_file_path()
        
        stats_text = f"""
📊 *Consultation Statistics*

Total consultations: {count}
Excel file: {excel_path.name}
Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        """
        
        await update.message.reply_text(stats_text, parse_mode='Markdown')
        
    except Exception as e:
        logger.error(f"Error getting stats: {e}")
        await update.message.reply_text("❌ Error retrieving statistics.")


async def myid_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Show user and chat IDs for admin setup."""
    chat_id = str(update.effective_chat.id)
    user_id = str(update.effective_user.id)
    user = update.effective_user
    chat = update.effective_chat
    
    # Determine chat type
    chat_type = "Private" if chat.type == 'private' else f"Group ({chat.type})"
    
    id_text = f"""
🆔 *Your Information*

**User ID:** `{user_id}`
**Chat ID:** `{chat_id}`
**Chat Type:** {chat_type}
**Username:** @{user.username if user.username else 'Not set'}
**Name:** {user.full_name}

*For admin setup, use your User ID:*
`ADMIN_CHAT_ID={user_id}`

*Current admin ID:* `{ADMIN_CHAT_ID if ADMIN_CHAT_ID else 'Not set'}`
    """
    
    await update.message.reply_text(id_text, parse_mode='Markdown')


async def handle_consultation_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle incoming consultation request messages."""
    try:
        message_text = update.message.text
        chat_id = str(update.effective_chat.id)
        
        # Parse the consultation data
        consultation_data = parse_consultation_message(message_text)
        
        if not consultation_data:
            await update.message.reply_text(
                "❌ Unable to parse consultation request. Please check the format.\n\n"
                "Required format:\n"
                "New Consultation Request\n"
                "Name: Your Name\n"
                "Email: your@email.com\n"
                "Phone: Your Phone\n"
                "Message: Optional message\n"
                "Date: Optional date\n\n"
                "Use /help to see more details."
            )
            return
        
        # Add chat_id to consultation data
        consultation_data['chat_id'] = chat_id
        
        # Add to Excel
        logger.info(f"Processing consultation from chat {chat_id}: {consultation_data.get('name', 'Unknown')}")
        success = consultation_manager.add_consultation(consultation_data)
        
        if success:
            logger.info(f"Successfully added consultation to Excel")
            # Send confirmation
            name = consultation_data.get('name', 'Unknown')
            response_text = f"✅ Consultation request received!\n\n"
            response_text += f"👤 Patient: {name}\n"
            response_text += f"📞 Phone: {consultation_data.get('phone', 'Not provided')}\n"
            response_text += f"📧 Email: {consultation_data.get('email', 'Not provided')}\n"
            
            # Add optional fields if they exist
            if consultation_data.get('message'):
                response_text += f"💬 Message: {consultation_data.get('message')}\n"
            if consultation_data.get('date'):
                response_text += f"📅 Date: {consultation_data.get('date')}\n"
            
            response_text += f"\nYour request has been saved and our team will contact you soon!"
            
            await update.message.reply_text(response_text)
            
            # Notify admin if configured
            if ADMIN_CHAT_ID and ADMIN_CHAT_ID != chat_id:
                admin_notification = f"🔔 *New Consultation Request*\n\n"
                admin_notification += f"👤 *Name:* {name}\n"
                admin_notification += f"📞 *Phone:* {consultation_data.get('phone', 'Not provided')}\n"
                admin_notification += f"📧 *Email:* {consultation_data.get('email', 'Not provided')}\n"
                
                # Add optional fields if they exist
                if consultation_data.get('message'):
                    admin_notification += f"💬 *Message:* {consultation_data.get('message')}\n"
                if consultation_data.get('date'):
                    admin_notification += f"� *Date:* {consultation_data.get('date')}\n"
                
                admin_notification += f"📱 *Chat ID:* {chat_id}\n"
                admin_notification += f"⏰ *Time:* {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                
                try:
                    await context.bot.send_message(
                        chat_id=ADMIN_CHAT_ID,
                        text=admin_notification,
                        parse_mode='Markdown'
                    )
                except Exception as e:
                    logger.error(f"Error sending admin notification: {e}")
            
        else:
            await update.message.reply_text(
                "❌ Error processing your consultation request. Please try again later."
            )
            
    except Exception as e:
        logger.error(f"Error handling consultation message: {e}")
        await update.message.reply_text(
            "❌ An error occurred while processing your request. Please try again."
        )


def main() -> None:
    """Start the bot."""
    if not BOT_TOKEN:
        logger.error("BOT_TOKEN not found in environment variables")
        return
    
    # Create the Application
    application = Application.builder().token(BOT_TOKEN).build()
    
    # Register handlers
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("help", help_command))
    application.add_handler(CommandHandler("myid", myid_command))
    application.add_handler(CommandHandler("get_consultations", get_consultations))
    application.add_handler(CommandHandler("stats", stats_command))
    
    # Handle all text messages as potential consultation requests
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_consultation_message))
    
    # Determine if we're running on a cloud platform or locally
    if WEBHOOK_URL:
        # Cloud hosting mode with webhooks
        logger.info("Starting AuraDent Bot in webhook mode...")
        application.run_webhook(
            listen="0.0.0.0",
            port=PORT,
            webhook_url=f"{WEBHOOK_URL}{WEBHOOK_PATH}",
            url_path=WEBHOOK_PATH,
            allowed_updates=Update.ALL_TYPES
        )
    else:
        # Local development mode with polling
        logger.info("Starting AuraDent Bot in polling mode...")
        application.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == '__main__':
    main()